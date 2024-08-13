"""Script to generate the performance report"""
import numpy as np
import pandas as pd
from os.path import join
from collections import OrderedDict
from sklearn.metrics import accuracy_score, recall_score, precision_score, roc_auc_score, confusion_matrix, f1_score
from openpyxl.drawing.image import Image
from openpyxl import Workbook
import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt
import eli5
from eli5.sklearn import PermutationImportance

from ml.config_train import CONFIG_TR
from utils import write_in_excel_df

def proba_to_int(y, thresh):
    """Apply threshold to probabilities"""
    return np.array(y > thresh, dtype=int)


def ws_permutation_importance(X_test, y_test, trained_model, wb):
    """Create worksheet with feature permutation importance results"""
    ws = wb.create_sheet("Permutation_Importance")

    perm = PermutationImportance(trained_model, random_state=1).fit(np.array(X_test), y_test)
    df_feat = eli5.explain_weights_df(perm, feature_names=X_test.columns.tolist())

    # feature        weight       std
    write_in_excel_df(ws, df_feat)

    fig = plt.figure()
    df_feat.plot(kind='barh', x='feature', y='weight', yerr="std")
    plt.gcf().set_size_inches(14, 4.5)

    path_temp = join(CONFIG_TR["dir_output"], "permutation_importances.png")
    plt.savefig(path_temp)
    img1 = Image(path_temp)
    ws.add_image(img1, "D1")
    plt.close()

    ordered_columns = list(df_feat["feature"])
    return ordered_columns


def performance_binary_classification(y_true, y_pred, thresh, name):
    """Computes  usual binary classification metrics"""
    dict_res = OrderedDict()
    dict_res["name"] = name

    acc = accuracy_score(y_true, proba_to_int(y_pred, thresh))
    dict_res["Accuracy"] = acc
    prec = precision_score(y_true, proba_to_int(y_pred, thresh))
    dict_res["Precision"] = prec
    rec = recall_score(y_true, proba_to_int(y_pred, thresh))
    dict_res["Recall"] = rec
    f1 = f1_score(y_true, proba_to_int(y_pred, thresh))
    dict_res["f1 score"] = f1
    auc = roc_auc_score(y_true, y_pred)

    dict_res["AUC"] = auc
    cm = confusion_matrix(y_true, proba_to_int(y_pred, thresh))
    cm = pd.DataFrame(cm)
    dict_res["Confusion Matrix"] = cm

    print("\t{}".format(name))
    print("\tAccuracy: {}".format(acc))
    print("\tPrecision: {}".format(prec))
    print("\tRecall: {}".format(rec))
    print("\tf1 score: {}".format(f1))
    print("\tAUC: {}".format(auc))
    print("\tConfusion Matrix: \n{}".format(cm))
    return dict_res


def ws_metrics_clas(y_train, y_test, y_pred_tr, y_pred_te, wb):
    """ Create a worksheet with summary Metrics"""

    # Metrics
    thresh = CONFIG_TR["threshold_classification"]
    metrics_tr = performance_binary_classification(y_train, y_pred_tr, thresh, name="TRAIN")
    metrics_cv = performance_binary_classification(y_test, y_pred_te, thresh, name="TEST")

    ws = wb.create_sheet("Metrics")
    ws["A1"] = "metric"
    ws["B1"] = "value"
    i = 2
    for k in metrics_tr.keys():
        if k != "Confusion Matrix":
            ws["A" + str(i)] = k
            ws["B" + str(i)] = metrics_tr[k]
            i += 1
        else:
            write_in_excel_df(ws, metrics_tr[k], t_row=i)
            i += 4
    for k in metrics_cv.keys():
        if k != "Confusion Matrix":
            ws["A" + str(i)] = k
            ws["B" + str(i)] = metrics_cv[k]
            i += 1
        else:
            write_in_excel_df(ws, metrics_cv[k], t_row=i)
            i += 4


def ws_config(wb):
    """ Create a worksheet with running configuration parameters"""

    ws = wb.create_sheet("Config")
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    i = 2
    for k in sorted(list(CONFIG_TR.keys())):
        ws["A" + str(i)] = k
        ws["B" + str(i)] = CONFIG_TR[k]
        i += 1



def ws_basic_feature_importance(feature_columns, algo, wb):
    """Create a worksheet with the algo's inner feature importance"""
    # importance = algo.coef_
    ws = wb.create_sheet("Feature_importances")

    if hasattr(algo, "feature_importances_"):
        importance = algo.feature_importances_
    elif hasattr(algo, "coef_"):
        importance = algo.coef_[0]

    a = pd.DataFrame()
    a["feat"] = feature_columns
    a["importance"] = importance
    a = a.sort_values(by=["importance"], ascending=False)
    if len(a) > 20:
        a = pd.concat([a.head(10), a.tail(10)], axis=0)  # type:pd.DataFrame
    write_in_excel_df(ws, a)

    fig = plt.figure()
    a.plot(kind='barh', x='feat', y='importance')

    path_temp = join(CONFIG_TR["dir_output"], "feature_importances_basic.png")
    plt.savefig(path_temp)
    img1 = Image(path_temp)
    ws.add_image(img1, "D1")
    plt.close()


def ws_actual_vs_pred_clas(y_train, y_test, y_pred_tr, y_pred_te, wb):
    """Create a worksheet with a prediction vs actual bubble chart"""
    ws = wb.create_sheet("Actual_vs_Predicted")
    # bubble plot
    plt.figure()
    ax = plt.axis()
    plt.title("Prediction versus Actual")
    plt.xlabel("ACTUAL")
    plt.ylabel("PREDICTED")
    final_pred_cv = pd.concat([pd.Series(y_test, name="actual"), pd.Series(y_pred_te, name="pred")],
                              axis=1)  # type: pd.DataFrame
    final_pred_cv["count"] = 1
    final_pred_bubble = final_pred_cv.groupby(by=["actual", "pred"], as_index=False).agg({"count": "count"})
    # sns.lmplot("actual", "pred", data= final_pred_bubble, hue="Id", fit_reg=False)

    plt.scatter("actual", "pred", data=final_pred_bubble, s="count")
    # plt.scatter("actual", "pred", data=final_pred_cv)

    path_temp = join(CONFIG_TR["dir_output"], "actual_vs_predicted.png")
    plt.savefig(path_temp)
    img = Image(path_temp)
    ws.add_image(img, 'A1')
    plt.close()


def assess_performance(y_train, y_test, y_pred_tr, y_pred_te, algo, feature_columns, X_test, y_test_cv):
    """ Create and save an Excel with a performance report"""
    # creating excel reviews
    wb = Workbook()
    columns_by_importance = None

    # Config
    ws_config(wb)

    # Metrics
    ws_metrics_clas(y_train, y_test, y_pred_tr, y_pred_te, wb)

    # actu vs pred
    ws_actual_vs_pred_clas(y_train, y_test, y_pred_tr, y_pred_te, wb)

    # # permutation importance
    columns_by_importance = ws_permutation_importance(X_test, y_test_cv, algo, wb)

    exc_name = join(CONFIG_TR["dir_output"], "performance_report_{}_{}.xlsx".format(CONFIG_TR["scenario"], CONFIG_TR["TIMESTAMP"]))
    wb.save(exc_name)
    print("\t\tPerformance report saved")
