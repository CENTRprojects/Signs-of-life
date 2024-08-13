"""Script gathering Excel related utils functions"""
from openpyxl.drawing.image import Image

from ml.retraining.imports import *
from ml.retraining.config_retrain import *
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Color, PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

import matplotlib as mpl

from ml.retraining.utils import gtm

mpl.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, Normalize, ListedColormap

GRAPH_ROW_HEIGHT = 30
PERC_COL_WIDTH = 7
LAB_COL_WIDTH = 70
URL_COL_WIDTH = 20

# COlOR SCALE
MIN_TH = -70
MAX_TH = 70


def get_smap():
    """Linear color gradient between MIN_TH and MAX_TH"""
    cmap = LinearSegmentedColormap.from_list('rg', ["r", "w", "g"], N=256)
    norm = Normalize(vmin=MIN_TH, vmax=MAX_TH)
    return plt.cm.ScalarMappable(cmap=cmap, norm=norm)


SM = get_smap()


def get_sm_color(value):
    """Converts a value into a HEX color using a linear gradient"""
    try:
        r, g, b = SM.to_rgba(value, bytes=True)[0:3]
    except:  # empty values(like total row)
        r = g = b = 255
    return rgb2hex(int(r), int(g), int(b))


def rgb2hex(r, g, b):
    return "{:02x}{:02x}{:02x}".format(r, g, b)


OPTIN_COLORS = {
    -1: rgb2hex(255, 160, 122),
    "no": rgb2hex(211, 211, 211),
    "yes": rgb2hex(173, 255, 47),
    0: rgb2hex(211, 211, 211),
    1: rgb2hex(173, 255, 47),
    2: rgb2hex(0, 150, 0)
}

OPTIN_COUNT_COLORS = {
    0: rgb2hex(173, 255, 47),
    1: rgb2hex(255, 255, 0),
    2: rgb2hex(255, 160, 122)
}

MAIN_COLOR = rgb2hex(152, 251, 152)
COLOR_HIGHLIGHT = rgb2hex(152, 251, 152)
COLOR_KEY = rgb2hex(183, 222, 232)
COLOR_IMPORTANT = rgb2hex(255, 197, 0)
COLOR_OTHER = rgb2hex(240, 240, 240)

# OTHER FORMATTING VARS
MAX_CELL_WIDTH = 40
MIN_CELL_WIDTH = 8

xl_border = Border(left=Side(border_style='thin', color='000000'),
                   right=Side(border_style='thin', color='000000'),
                   top=Side(border_style='thin', color='000000'),
                   bottom=Side(border_style='thin', color='000000'))
FORMAT_PERCENT = '0.00%'
FORMAT_GENERAL = 'General'

# Sheets names
DICO_FUNC_TO_NAMES = {
    "add_ws_config": "CONFIG",
    "add_ws_log": "LOG",
    "add_ws_labeller": "Labeller",
    "add_ws_performance": "Performance",
    "add_ws_history": "history",
    "add_ws_error": "error_analysis",
    "add_ws_detail": "details",
}

# user messages
UPSTREAM_ERROR_MSG = "Please Check LOG. Not available yet due to upstream error."


def color_cells_with_sm(row, col, ws, idx=""):
    """Apply color fill for one cell from its value through a linear gradient color"""
    value = ws[get_column_letter(col) + str(row)].value
    try:
        r, g, b = SM.to_rgba(value, bytes=True)[0:3]
        ws[get_column_letter(col) + str(row)].fill = PatternFill("solid", fgColor=rgb2hex(int(r), int(g), int(b)))
    except Exception as e:
        pass
        # add_to_log("Unexpected value {}:{}".format(value, idx))


def write_in_excel_df_raw(ws, df, t_col=0, t_row=0, with_header=True):
    """Write a dataframe in an Excel sheet"""
    # add_to_log("EXCL columns: \t\t\t{}".format(",".join(list(df.columns))))
    rows = dataframe_to_rows(df, index=False, header=with_header)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + t_row, column=c_idx + t_col, value=value)


def location_to_range(df, t_col, t_row, with_header=True):
    """Identifies the Excel range or coordinates (ex: A1:B3) of a dataframe"""
    rows = df.shape[0]
    cols = df.shape[1]

    r_min = t_row + 1
    if with_header:
        r_max = t_row + 1 + rows
    else:
        r_max = t_row + rows
    c_min = get_column_letter(t_col + 1)
    c_max = get_column_letter(t_col + cols)

    return c_min + str(r_min) + ":" + c_max + str(r_max)


def write_in_excel_df(ws, df, t_col=0, t_row=0, with_header=True, dico_fmts=None, dico_user_rename=None):
    """Write a dataframe into Excel with columns renaming and columns/value formatting options"""
    apply_formating = dico_fmts is not None
    dico_idx_to_col = dict(zip(list(range(len(list(df.columns)))), list(df.columns)))

    rows = dataframe_to_rows(df, index=False, header=with_header)

    # width
    for icol in range(df.shape[1]):
        if apply_formating:
            if dico_idx_to_col[icol] in dico_fmts:
                dico_colo_fmt = dico_fmts[dico_idx_to_col[icol]]
            else:
                dico_colo_fmt = dico_fmts["X_OTHER"]

            width = dico_colo_fmt["width"]
            cur_width = ws.column_dimensions[get_column_letter(t_col + icol + 1)].width
            if width == "auto":
                colo_name = dico_idx_to_col[icol]
                if len(df) > 0:
                    max_lg = max([df[colo_name].apply(lambda x: len(str(x))).max(), len(str(colo_name))])
                else:
                    max_lg = len(str(colo_name))
                colo_width = max([min([int(1.1 * max_lg), MAX_CELL_WIDTH]), MIN_CELL_WIDTH])
                ws.column_dimensions[get_column_letter(t_col + icol + 1)].width = max([cur_width, colo_width])
            elif isinstance(width, float) or isinstance(width, int):
                ws.column_dimensions[get_column_letter(t_col + icol + 1)].width = max([cur_width, width])

    # header and cols
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cl = ws.cell(row=r_idx + t_row, column=c_idx + t_col)

            # border
            cl.border = xl_border

            is_header = r_idx == int(with_header)
            if not apply_formating:
                if not is_header:
                    cl.value = value
                else:

                    # user rename
                    final_column_name = value
                    if with_header and (dico_user_rename is not None):
                        if value in dico_user_rename:
                            # write renamed_columns
                            final_column_name = dico_user_rename[value]
                    cl.value = final_column_name

            else:

                # fetch column format
                colo_name = dico_idx_to_col[c_idx - 1]
                if colo_name in dico_fmts:
                    dico_colo_fmt = dico_fmts[colo_name]
                else:
                    dico_colo_fmt = dico_fmts["X_OTHER"]

                if is_header:

                    # color
                    if dico_colo_fmt["type"] == "key":
                        cl.fill = PatternFill("solid", fgColor=COLOR_KEY)
                    elif dico_colo_fmt["type"] == "important":
                        cl.fill = PatternFill("solid", fgColor=COLOR_IMPORTANT)
                    else:
                        cl.fill = PatternFill("solid", fgColor=COLOR_OTHER)

                    # align
                    if dico_colo_fmt["align"] == "center":
                        cl.alignment = Alignment(horizontal='center')
                    elif dico_colo_fmt["align"] == "left":
                        cl.alignment = Alignment(horizontal='left')

                    # user rename
                    final_column_name = value
                    if with_header and (dico_user_rename is not None):
                        if value in dico_user_rename:
                            # write renamed_columns
                            final_column_name = dico_user_rename[value]

                    # set value
                    cl.value = final_column_name

                else:

                    # align
                    if dico_colo_fmt["align_value"] == "center":
                        cl.alignment = Alignment(horizontal='center')
                    elif dico_colo_fmt["align_value"] == "left":
                        cl.alignment = Alignment(horizontal='left')

                    # functional_colo
                    if dico_colo_fmt["is_func_colo"]:
                        clr = dico_colo_fmt["func_color"]
                        if clr == "sm_percent":
                            # add_to_log(get_sm_color(value))
                            cl.fill = PatternFill("solid", fgColor=get_sm_color(value))
                        elif clr == "main":
                            cl.fill = PatternFill("solid", fgColor=COLOR_HIGHLIGHT)
                        else:
                            cl.fill = PatternFill("solid", fgColor=clr)

                    # format
                    try:
                        final_value = value
                        if dico_colo_fmt["format_value"] is not None:
                            tp = dico_colo_fmt["format_value"]

                            if tp == FORMAT_PERCENT:
                                # percentage adaptation
                                final_value = value / 100

                            cl.number_format = tp

                        # set value
                        cl.value = final_value

                    except:  # empty cells (total row...)
                        cl.value = value

    # filter
    filter_range = location_to_range(df, t_col, t_row)
    ws.auto_filter.ref = filter_range


def add_trivial_sheet(wb, sh_name, title, err_mess):
    """Creates a basic sheet, if not existing yet, with an error message"""
    if sh_name not in wb.sheetnames:
        ws = wb.create_sheet(sh_name)
    else:
        ws = wb[sh_name]
    ws.sheet_properties.tabColor = "000000"
    ws["B1"] = title
    ws["C1"] = err_mess


def exc_handler_with_wb(func):
    """Wraps a sheet-creating function with a sheet-creating Exception catch clause"""

    def exc_run(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            nm = func.__name__
            sh_name = DICO_FUNC_TO_NAMES[nm]
            err_mess = "{} \n {} \n {}".format(type(e), str(e), traceback.format_exc())
            wb = None
            for elem in args:
                if isinstance(elem, openpyxl.workbook.Workbook):
                    wb = elem
            if wb is None:
                add_to_log("ERROR: The decorator exc_handler_with_wb is only reserved to functions with a "
                           "workbook as one of the argument")
            else:
                add_trivial_sheet(wb, sh_name, sh_name, err_mess)

    return exc_run


def fill_spe_range(ws, min_row, max_row, min_col, max_col, color=COLOR_OTHER):
    """Fill a range of Excel cells with the provided color"""
    for c in range(min_col, max_col + 1):
        for r in range(min_row, max_row + 1):
            ws[get_column_letter(c) + str(r)].fill = PatternFill("solid", fgColor=color)


def write_format_one_cell(ws, row, col, value, tp):
    """Format the Excel cell with a Value formatting option (percent, dollar..)"""
    if tp is not None:
        if tp == FORMAT_PERCENT:
            # percentage adaptation
            try:
                new_value = value / 100
                cl = ws.cell(row=row, column=col, value=new_value)
                cl.number_format = tp
            except:  # empty values (like total row)
                cl = ws.cell(row=row, column=col, value=value)
        else:
            cl = ws.cell(row=row, column=col, value=value)
            cl.number_format = tp


def wrap_text_one_cell(ws, row, col):
    """Apply the Excel function Wrap Text at a given cell"""
    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)


def freeze_pane(ws, cell_address):
    """Apply the Excel function Wrap Text at a given cell"""
    try:
        cell = ws[cell_address]
        ws.freeze_panes = cell
    except Exception as e:
        add_to_log("EXCEL freeze pane error --> ignored: {} - {} - {}".format(type(e), str(e), traceback.format_exc()))
        pass


@exc_handler_with_wb
def add_ws_log(wb):
    """Creates, derive data and format the sheet LOG"""
    ws = wb.create_sheet(DICO_FUNC_TO_NAMES["add_ws_log"])
    # ws.sheet_properties.tabColor = "28D6FB"

    ws["A1"] = "LOG (=Warnings, Errors and Critical Errors occuring during calculation)"

    if LG.logs:

        LG.add({"time": gtm(), "importance": "INFO", "category": "formatting", "description": "end"})

        df_log = pd.DataFrame(LG.logs)

        for colo in LG.expected_cols:
            if colo not in df_log.columns:
                df_log[colo] = None

        df_log = df_log[LG.expected_cols]
        write_in_excel_df(ws, df_log.fillna("").astype(str), t_row=1)

    else:
        ws["A2"] = "NO LOG"


# @exc_handler_with_wb
def add_ws_labeller(wb, input_data_train):
    """Creates, derive data and format the sheet LOG"""
    ws = wb.create_sheet(DICO_FUNC_TO_NAMES["add_ws_labeller"])
    # ws.sheet_properties.tabColor = "28D6FB"

    ws["A1"] = "Labeller analysis"

    row = 2

    tables = [("summary_agreement", "Agreement summary"),
              # ("disagreement", "Disagreements"),
              ("top_disagreement", "Top disagreements samples"),
              ("labeller_view", "Labeller analysis")]

    for tb_name, tb_title in tables:

        ws[f"A{row}"] = tb_title
        if tb_name in input_data_train:
            df = input_data_train[tb_name]

            if tb_name == "labeller_view":
                # ordering
                colos_order = ["labeller", "QTY_labels", "QTY_labels_with_multi_lbls", "ratio_full_agreement",
                               "ratio_one_agreement"]
                df = df[colos_order]

            if tb_name == "top_disagreement":
                from openpyxl.formatting.rule import DataBarRule
                from openpyxl.styles import colors

                col_min = 5
                cur_row = row + 2
                for j, row_df in df.iterrows():
                    labels = row_df["label_counts"]
                    n_labels = row_df["n_labels"]
                    cur_col = col_min

                    # compute percentages
                    labels_percs_rnd = compute_label_percentage(labels)

                    # color percentage cells
                    for i_lab, label in enumerate(labels_percs_rnd):
                        cnt_label = int(label[1] / 10)

                        if label[0] in DICO_COLORS_TO_LV2:
                            clr = DICO_COLORS_TO_LV2[label[0]]
                        else:
                            print(label[0])
                            clr = DEFAULT_COL

                        for i_col in range(cur_col, cur_col + cnt_label):
                            cell_addr = get_column_letter(i_col) + str(cur_row)
                            ws[cell_addr].fill = PatternFill("solid", fgColor=clr)
                        ws.cell(row=cur_row, column=cur_col).value = label[1]

                        cur_col += cnt_label

                    cur_row += 1
                    a = 1

                # width
                for colo_perc in range(col_min, col_min + 10):
                    ws.column_dimensions[get_column_letter(colo_perc)].width = PERC_COL_WIDTH

            write_in_excel_df(ws, df, t_row=row)

            if tb_name == "top_disagreement":
                # add screenshot hyperlinks
                cur_row = row + 2
                for j, row_df in df.iterrows():
                    fp_fold = join(MAIN_DIR, SCREENSHOT_FOLDER, row_df["ss_filename"])

                    ws['B' + str(cur_row)].value = f'=HYPERLINK("{fp_fold}", "screenshot")'
                    ws['B' + str(cur_row)].style = 'Hyperlink'
                    cur_row += 1

            row += len(df) + 3

        else:
            ws[f"A{row + 1}"] = f"Table {tb_name} not found"
            row += 3

    ws.column_dimensions["A"].width = URL_COL_WIDTH
    ws.column_dimensions["D"].width = LAB_COL_WIDTH


def compute_label_percentage(labels):
    labels_counts = [e.split(":") for e in labels.split(",")]
    labels_counts = [[e[0].strip(), float(e[1].strip())] for e in labels_counts]
    tot_labels = sum([e[1] for e in labels_counts])
    labels_percs = [[e[0], int(100 * e[1] / tot_labels)] for e in labels_counts]
    labels_percs_rnd = [[e[0], int(e[1] / 10) * 10] for e in labels_percs]
    sum_perc = sum([e[1] for e in labels_percs_rnd])
    if sum_perc != 100:
        maxi = 0
        i_maxi = 0
        for i_cur, pc in enumerate(labels_percs_rnd):
            diff = (pc[1] - 10 * int(pc[1] / 10))
            if diff > 0:
                if diff > maxi:
                    i_maxi = i_cur
                    maxi = diff
        labels_percs_rnd[i_maxi][1] += 10
    return labels_percs_rnd


@exc_handler_with_wb
def add_ws_error(wb, output_data):
    """Creates, derive data and format the sheet LOG"""
    ws = wb.create_sheet(DICO_FUNC_TO_NAMES["add_ws_error"])
    # ws.sheet_properties.tabColor = "28D6FB"

    ws["A1"] = "Error analysis"

    row = 2

    tb_name = "prediction_errors"
    if tb_name in output_data:
        df = output_data[tb_name]

        # to remove
        colo_rem = ["html_raw", "html_text", "raw_text"]
        for colo in colo_rem:
            if colo in df.columns:
                df = df.drop([colo], axis=1)

        # reorder
        first_cols = ["url", "target", "predicted", "ss_filename", "target_label", "suggested_error", "n_words",
                      "n_words_attrib"]
        for colo in first_cols:  # complete missing
            if colo not in list(df.columns):
                df[colo] = np.nan
        all_cols = list(df.columns)
        other_cols = [e for e in all_cols if e not in first_cols]
        df = df[first_cols + other_cols]

        df = df.astype(str)

        write_in_excel_df(ws, df, t_row=row)

        # add screenshot hyperlinks
        cur_row = row + 2
        colo_screenshot = "D"
        for j, row_df in df.iterrows():
            fp_fold = join(MAIN_DIR, SCREENSHOT_FOLDER, row_df["ss_filename"])

            ws[colo_screenshot + str(cur_row)].value = f'=HYPERLINK("{fp_fold}", "screenshot")'
            ws[colo_screenshot + str(cur_row)].style = 'Hyperlink'
            cur_row += 1

        row += len(df) + 3

    else:
        ws[f"A{row + 1}"] = f"Table {tb_name} not found"
        row += 3


@exc_handler_with_wb
def add_ws_detail(wb, output_data):
    """Creates, derive data and format the sheet LOG"""
    ws = wb.create_sheet(DICO_FUNC_TO_NAMES["add_ws_detail"])
    # ws.sheet_properties.tabColor = "28D6FB"

    ws["A1"] = "All predictions"

    row = 2

    tb_name = "all_predictions"
    if tb_name in output_data:
        df = output_data[tb_name]

        first_cols = ["url", "target", "predicted", "ss_filename"]
        all_cols = list(df.columns)
        other_cols = [e for e in all_cols if e not in first_cols]
        df = df[first_cols + other_cols]

        # to remove
        colo_rem = ["html_raw", "html_text", "raw_text"]
        for colo in colo_rem:
            if colo in df.columns:
                df = df.drop([colo], axis=1)

        # ILLEGAL_CHARACTERS_RE.sub(r'', x)

        df = df.astype(str)

        write_in_excel_df(ws, df, t_row=row)

        # add screenshot hyperlinks
        cur_row = row + 2
        colo_screenshot = "D"
        for j, row_df in df.iterrows():
            fp_fold = join(MAIN_DIR, SCREENSHOT_FOLDER, row_df["ss_filename"])

            ws[colo_screenshot + str(cur_row)].value = f'=HYPERLINK("{fp_fold}", "screenshot")'
            ws[colo_screenshot + str(cur_row)].style = 'Hyperlink'
            cur_row += 1

        row += len(df) + 3

    else:
        ws[f"A{row + 1}"] = f"Table {tb_name} not found"
        row += 3


@exc_handler_with_wb
def add_ws_config(wb):
    """Creates, derive data and format the sheet LOG"""
    ws = wb.create_sheet(DICO_FUNC_TO_NAMES["add_ws_config"])
    # ws.sheet_properties.tabColor = "28D6FB"

    ws["A1"] = "Retraining configuration"

    row = 2

    nm = "Parameter Name"
    vl = "Parameter Value"
    dc = {nm: "", vl: ""}

    # config
    PMS = ["VERSION_SELECTION", "N_TOP_DISAGREEMENT", "N_MAX_LABELS", "MIN_PREVAILING_THRESH",
           "ML_DO_PERFS", "TRAIN_AT_LV2", "CROSS_VAL_N_SPLIT", "THRESHOLD_CLASSIFICATION", "DO_FULL_TR",
           "VERSION_TOKENISATION", "VERSION_TRAINING"]
    PMS_DICO = ["CONFIG_FENG", "CONFIG_XGB", "CONFIG_XGB_MCLASS"]

    lst = []
    for pm in PMS:
        dc_cur = copy.deepcopy(dc)
        dc_cur[nm] = str(pm)
        dc_cur[vl] = str(eval(pm))
        lst.append(dc_cur)

    for pm in PMS_DICO:
        # param name
        dc_cur = copy.deepcopy(dc)
        dc_cur[nm] = pm
        dc_cur[vl] = ""
        lst.append(dc_cur)

        # values
        sub_pm = sorted([(str(k), v) for k, v in eval(pm).items()], key=lambda x: x[0])
        for key, value in sub_pm:
            dc_cur = copy.deepcopy(dc)
            dc_cur[nm] = str(key)
            dc_cur[vl] = str(value)
            lst.append(dc_cur)

    df = pd.DataFrame(lst)
    write_in_excel_df(ws, df, t_row=row)
    row += len(df) + 5

    # add label categorys
    if not TRAIN_AT_LV2:
        ws[f"A{row}"] = "Caption number to labels LV3"
        row += 1
        lbls = []
        for k, v in DICO_LABELS_TO_LV3.items():
            lbls.append([v, k])

        lbls = sorted(lbls, key=lambda x: x[0])
        for tpl in lbls:
            ws[f"A{row}"] = tpl[0]
            ws[f"B{row}"] = tpl[1]
            row += 1


# @exc_handler_with_wb
def get_current_result(output_data, ref_time):
    dico_perf = output_data["CV_performance"]
    current_result = {"Accuracy": dico_perf["Accuracy"], "n_samples": dico_perf["n_samples"], "ref_time": ref_time}

    # todo: add incertitude range

    a = 1
    return current_result


def compute_general_accuracy(acc_ml, tld):
    if tld == "cctld":
        preval_err = PREVALENCE_ERRORS_CCTLD
        preval_regist = PREVALENCE_REGISTRARS_CCTLD
    elif tld == "gtld":
        preval_err = PREVALENCE_ERRORS_GTLD
        preval_regist = PREVALENCE_REGISTRARS_GTLD
    else:
        raise ValueError("unexpected TLD")

    preval_ml = 1 - preval_err - preval_regist
    acc_general = preval_ml * acc_ml + preval_err * ACCURACY_ERRORS + preval_regist * ACCURACY_REGISTRARS

    return acc_general


def plot_double_line(x, y1, y2, plt_name, plt_folder, metric_name):
    path_line = join(plt_folder, plt_name + ".png")

    # plt.figure()
    fig, ax = plt.subplots()
    plt.title(plt_name)
    ax.plot(x, y1, color="blue", marker="o")
    ax.set_xlabel("thresh_val", fontsize=14)
    ax.set_ylabel("metric= {}".format(metric_name), color="blue", fontsize=14)
    plt.xticks(rotation=20)

    # add qtty
    ax2 = ax.twinx()
    ax2.plot(x, y2, color="red", marker="o")
    ax2.set_ylabel("n_samples", color="red", fontsize=14)

    # size
    deft_size = fig.get_size_inches()
    wid_size = (int(1.25 * deft_size[0]), deft_size[1])
    fig.set_size_inches(wid_size)

    # close
    plt.savefig(path_line)
    plt.close()

    return path_line


def add_ws_history(wb, input_data, output_data, ref_time, out_folder):
    if "CV_performance" not in output_data:
        return None

    ws = wb.create_sheet(DICO_FUNC_TO_NAMES["add_ws_history"])
    LEGACY_REF_TIME = "20200625-165313"

    if "historical_perfs" not in input_data:
        historical_perfs = []
    else:
        historical_perfs = input_data["historical_perfs"]

    current_result = get_current_result(output_data, ref_time)
    historical_perfs.append(current_result)

    ws["A1"] = "Historical performance"

    for resu in historical_perfs:
        if resu["ref_time"] == LEGACY_REF_TIME:  # = legacy accuracy
            resu.update({"Accuracy_CCTLD": resu["Accuracy"]})
            resu.update({"Accuracy_GTLD": resu["Accuracy"]})
        else:
            resu.update({"Accuracy_CCTLD": compute_general_accuracy(resu["Accuracy"], tld="cctld")})
            resu.update({"Accuracy_GTLD": compute_general_accuracy(resu["Accuracy"], tld="gtld")})

    x = [e["ref_time"] for e in historical_perfs]
    y2 = [e["n_samples"] for e in historical_perfs]
    y1_ml_only = [e["Accuracy"] for e in historical_perfs]
    y1_full_cctld = [e["Accuracy_CCTLD"] for e in historical_perfs]
    y1_full_gtld = [e["Accuracy_GTLD"] for e in historical_perfs]

    col_img = 2
    row_img = 2

    # ML only
    plt_name = "history_ML_only"
    metric_name = "Accuracy (ML domains only)"
    path_fig_ml_only = plot_double_line(x, y1_ml_only, y2, plt_name, out_folder, metric_name)
    img = Image(path_fig_ml_only)
    col_img_letter = get_column_letter(col_img)
    ws.add_image(img, '{}{}'.format(col_img_letter, row_img))
    row_img += GRAPH_ROW_HEIGHT

    # CCTLD
    plt_name = "history_CCTLD"
    metric_name = "Accuracy (CCTLD)"
    path_fig_cctld = plot_double_line(x, y1_full_cctld, y2, plt_name, out_folder, metric_name)
    img = Image(path_fig_cctld)
    col_img_letter = get_column_letter(col_img)
    ws.add_image(img, '{}{}'.format(col_img_letter, row_img))
    row_img += GRAPH_ROW_HEIGHT

    # GTLD
    plt_name = "history_GTLD"
    metric_name = "Accuracy (GTLD)"
    path_fig_gtld = plot_double_line(x, y1_full_gtld, y2, plt_name, out_folder, metric_name)
    img = Image(path_fig_gtld)
    col_img_letter = get_column_letter(col_img)
    ws.add_image(img, '{}{}'.format(col_img_letter, row_img))

    return historical_perfs


@exc_handler_with_wb
def add_ws_performance(wb, output_data):
    """Creates, derive data and format the sheet LOG"""
    ws = wb.create_sheet(DICO_FUNC_TO_NAMES["add_ws_performance"])
    # ws.sheet_properties.tabColor = "28D6FB"

    ws["A1"] = "ML Performance"

    row = 2

    tables = [
        ("CV_performance", "Crossvalidation Data"), ("TR_performance", "Training Data")
    ]
    if TRAIN_AT_LV2:
        tables += [("performance_by_lgg", "Performance by language"), ("performance_by_tld", "Performance by TLD")]

    for tb_name, tb_title in tables:

        ws[f"A{row}"] = tb_title
        if tb_name in output_data:
            elem = output_data[tb_name]

            if isinstance(elem, OrderedDict):
                for k, v in elem.items():
                    if k != "Confusion Matrix":
                        ws[f"A{row + 1}"] = k
                        ws[f"B{row + 1}"] = str(v)
                        row += 1
                    else:
                        ws[f"A{row + 1}"] = k
                        if isinstance(v, pd.DataFrame):
                            write_in_excel_df(ws, v.astype(str), t_col=1, t_row=row)
                            row += len(v) + 1
                        else:
                            ws[f"B{row + 1}"] = str(v)
                            row += 1
                row += 3

            elif isinstance(elem, pd.DataFrame):
                df = elem
                write_in_excel_df(ws, elem, t_row=row)
                row += len(df) + 3
            else:
                ws[f"A{row + 1}"] = f"Unexpected format for table {tb_name}: {type(elem)}"
                row += 3

        else:
            ws[f"A{row + 1}"] = f"Table {tb_name} not found"
            row += 3
